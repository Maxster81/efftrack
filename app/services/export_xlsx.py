"""Servizio di esportazione XLSX per i record di effort (v1.9.0).

Genera un file `.xlsx` in memoria (senza scrivere su disco) usando `openpyxl`,
con le stesse colonne e la stessa logica dell'export CSV (`app/services/export_csv.py`):
- colonne identiche a `CSV_HEADER` (Data, Cliente, Gruppo, Attività, Utente, Ore, Note, Descrizione attività);
- esclusione dei giorni non lavorati (record sentinella "NON LAVORATO", S6);
- data formattata `gg/mm/aaaa`, ore come numero, header in grassetto e larghezza colonne calibrata.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.models import EffortEntry
from app.services.export_csv import CSV_HEADER, is_sentinel_entry

# Media type corretto per i file XLSX (Office Open XML Spreadsheet).
XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Larghezza (in unità di carattere) delle colonne, allineate a CSV_HEADER.
_COLUMN_WIDTHS = [12, 22, 14, 26, 28, 8, 32, 32]


def build_xlsx(records: list[EffortEntry]) -> bytes:
    """Genera un workbook `.xlsx` in memoria con i record di effort.

    I giorni non lavorati (record con cliente sentinella "NON LAVORATO", S6)
    vengono esclusi, come nell'export CSV. La colonna Utente mostra lo username
    reale dal JOIN su `users` (vuota per i record legacy senza proprietario).
    Restituisce i byte del file `.xlsx` (pronti per un download via HTTP).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Effort"

    ws.append(CSV_HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for idx, width in enumerate(_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for record in records:
        if is_sentinel_entry(record):
            # I giorni non lavorati non compaiono nell'export (S6).
            continue
        ws.append(
            [
                record.work_date.strftime("%d/%m/%Y"),
                record.client.name,
                record.group.name,
                record.activity.name,
                record.user.username if record.user is not None else "",
                # Ore come numero (più comodo in Excel per eventuali somme).
                float(record.hours_spent) if record.hours_spent is not None else None,
                record.notes or "",
                record.description or "",
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
